# -*- coding: utf-8 -*-
"""
Script para automação de criação de issues no Jira a partir de uma planilha Excel.

Este script lê uma planilha com detalhes de projetos, extrai as informações
necessárias, e cria automaticamente Épicos, Estórias e Tarefas no Jira.
Após a criação, ele atualiza a planilha com os IDs dos tickets gerados.

Pré-requisitos:
- Python 3.x
- Bibliotecas: pandas, openpyxl, numpy, jiralib (customizada)
- Arquivo 'setup.json' com as credenciais e URL do Jira.
- Planilha Excel no formato esperado.

Como executar:
python processa_refatorado.py --nome "SuaPlanilha.xlsx"
"""

import argparse
import jiralib
import json
import logging
import sys
from typing import Dict, Any, List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# --- 1. CONFIGURAÇÕES E CONSTANTES GLOBAIS ---
# Definir constantes melhora a legibilidade e facilita a manutenção!
ABA_PLANILHA = "Planilha"
ABA_PARAMETROS = "Parâmetros"
SETUP_FILE = "setup.json"

# Colunas essenciais que esperamos encontrar na planilha.
COLUNAS_DESEJADAS = [
    "Projeto", "Data", "Request", "Task", "Módulo", "Tipo", "Usuário Projeto",
    "UUID PO", "UUID Recurso", "Épico", "Estória", "Tarefa", "Etapa",
    "Horas", "Pontos", "TicketE", "TicketS", "TicketT", "Prioridade", "Retorno",
    "US"
]

def parse_argumentos() -> str:
    """
    Analisa os argumentos da linha de comando para obter o nome do arquivo.
    É uma ótima prática isolar a lógica de parsing!
    """
    parser = argparse.ArgumentParser(
        description="Processa uma planilha Excel para criar issues no Jira."
    )
    parser.add_argument(
        '--nome',
        type=str,
        required=True,
        help='Nome do arquivo da planilha Excel a ser processada.'
    )
    args = parser.parse_args()
    return args.nome

def carregar_configuracoes_jira() -> None:
    """
    Carrega as configurações de conexão com o Jira a partir do arquivo setup.json.
    Isso mantém as credenciais e a URL fora do código, o que é mais seguro!
    """
    try:
        with open(SETUP_FILE, 'r', encoding='utf-8') as f:
            config = json.load(f)
            # Supondo que 'jiralib' seja um módulo que você tem.
            # Se não for, você precisará importar ou definir 'jiralib'.
            jiralib.url = config['url']
            jiralib.headers = config['headers']
            jiralib.add_output("✓ Configurações do Jira carregadas com sucesso!")
    except FileNotFoundError:
        jiralib.add_output(f"❌ Erro: Arquivo de configuração '{SETUP_FILE}' não encontrado.")
        sys.exit(1) # Encerra o script se o arquivo de configuração não existir.
    except KeyError:
        jiralib.add_output(f"❌ Erro: O arquivo '{SETUP_FILE}' está mal formatado. 'url' ou 'headers' não encontrados.")
        sys.exit(1)

def preparar_dataframe(arquivo_excel: str) -> pd.DataFrame:
    """
    Lê a planilha e prepara o DataFrame para o processamento.
    Limpa os dados, preenche valores nulos e cria colunas customizadas.
    """
    jiralib.add_output("🚀 Iniciando a leitura e preparação dos dados da planilha...")
    
    # Usar um bloco try-except para ler o arquivo é uma boa prática para capturar erros.
    try:
        df = pd.read_excel(arquivo_excel, sheet_name=ABA_PLANILHA, engine="openpyxl")
    except FileNotFoundError:
        jiralib.add_output(f"❌ Erro: A planilha '{arquivo_excel}' não foi encontrada.")
        sys.exit(1)
    except ValueError as e:
        jiralib.add_output(f"❌ Erro ao ler a aba '{ABA_PLANILHA}': {e}")
        sys.exit(1)
        
    # Seleciona apenas as colunas que vamos usar para manter o DataFrame limpo.
    df = df[COLUNAS_DESEJADAS].copy()

    # Preenche valores NaN (nulos) para evitar erros na concatenação de strings.
    # Usar um dicionário torna o código mais limpo e organizado!
    valores_padrao = {
        "Épico": "", "Estória": "", "Request": "", "Task": "", "Prioridade": "0",
        "Tipo": "", "TicketE": "0", "TicketS": "0", "TicketT": "0", "Retorno": 0,
        "US": ""
    }
    df.fillna(valores_padrao, inplace=True)

    # Criação das colunas com nomes formatados para o Jira usando f-strings.
    # f-strings são mais rápidas e legíveis que a concatenação tradicional!
    df["Nome_Epico"] = df.apply(
        lambda row: f"[{row['Request']}] - {row['Épico']}", axis=1
    )
    df["Nome_Estoria"] = df.apply(
        lambda row: f"[{row['Tipo']}] - [{row['Task']}] - {row['Estória']}", axis=1
    )
    df["Usuario_Projeto"] = df.apply(
        lambda row: f"{row['Usuário Projeto']}", axis=1
    )
    df["UUID_PO"] = df.apply(
        lambda row: f"{row['UUID PO']}", axis=1
    )    
    df["UUID_Recurso"] = df.apply(
        lambda row: f"{row['UUID Recurso']}", axis=1
    )    

    jiralib.add_output("✓ DataFrame preparado com sucesso!")
    return df

def processar_issues_jira(df: pd.DataFrame) -> Dict[str, List[Any]]:
    """
    Itera sobre o DataFrame e cria Épicos, Estórias e Tarefas no Jira.
    Esta é a função principal que interage com a API do Jira!
    """

    # Listas para armazenar os tickets criados.
    epicos_criados = set()
    estorias_criadas = set()
    
    # Dicionário para armazenar os resultados.
    resultados = {
        "Ticket_Epico": [],
        "Ticket_Estoria": [],
        "Ticket_Task": []
    }

    last_epic_key = None
    last_story_key = None
    
    jiralib.add_output("\n🔥 Começando a integração com o Jira! Criando issues...")

    # Iterar com df.itertuples() é mais performático que to_dict.
    for row in df.itertuples():
        
        # --- CRIAÇÃO DO ÉPICO ---
        # A chave única para identificar se um épico já foi criado nesta execução.
        epic_key_identificador = f"[{row.Request}] - {row.Épico}"
        
        if row.TicketE == "0":
            if epic_key_identificador not in epicos_criados:
                jiralib.add_output(f"✨ Criando Épico: {row.Nome_Epico}"[:80]) # _19 é o índice de "Nome Épico"
                try:
                    response = jiralib.create_epic(
                        row.Projeto, row.Nome_Epico, row.Nome_Epico, row.Data, row.Nome_Epico,
                        row.Usuario_Projeto, row.Task, row.Módulo, row.Tipo, row.Prioridade, row.Retorno
                    )
                    response.raise_for_status() # Lança um erro para respostas HTTP 4xx/5xx
                    last_epic_key = response.json().get("key")
                    jiralib.add_output(f"    ✅ Sucesso! Chave do Épico: {last_epic_key}")
                except Exception as e:
                    jiralib.add_output(f"    ❌ Falha ao criar Épico. Erro: {e}")
                    last_epic_key = "ERRO"
                if last_epic_key != "ERRO":
                    #to: Ready for Development
                    response = jiralib.advance_status(last_epic_key, "11")
                    #to: To Do
                    response = jiralib.advance_status(last_epic_key, "21")
                    epicos_criados.add(epic_key_identificador)
            else:
                jiralib.add_output(f"⏭️  Pulando Épico já criado nesta execução: ({last_epic_key})")
        else:
            last_epic_key = row.TicketE
            jiralib.add_output(f"⏭️  Pulando Épico já existente: ({last_epic_key})")

        # --- CRIAÇÃO DA ESTÓRIA ---
        story_key_identificador = f"[{row.Task}] - {row.Estória}"

        if row.TicketS == "0":
            if story_key_identificador not in estorias_criadas:
                jiralib.add_output(f"  ✨ Criando Estória: {row.Nome_Estoria}"[:80]) # _20 é o índice de "Nome Estória"
                try:
                    response = jiralib.create_story(
                        row.Projeto, "História", row.Nome_Estoria, row.US, row.UUID_Recurso,
                        row.UUID_PO, last_epic_key, row.Pontos
                    )
                    response.raise_for_status()
                    last_story_key = response.json().get("key")
                    jiralib.add_output(f"      ✅ Sucesso! Chave da Estória: {last_story_key}")
                except Exception as e:
                    jiralib.add_output(f"      ❌ Falha ao criar Estória. Erro: {e}")
                    last_story_key = "ERRO"
                if last_epic_key != "ERRO":
                    #to: Defined
                    response = jiralib.advance_status(last_epic_key, "21")
                    estorias_criadas.add(story_key_identificador)
            else:
                jiralib.add_output(f"  ⏭️  Pulando Estória já criado nesta execução: ({last_story_key})")
        else:
            last_story_key = row.TicketS
            jiralib.add_output(f"  ⏭️  Pulando Estória já existente: ({last_story_key})")

        # --- CRIAÇÃO DA TAREFA ---
        if row.TicketT == "0":
            jiralib.add_output(f"    ✨ Criando Tarefa: {row.Tarefa}"[:80])
            try:
                response = jiralib.create_task(
                    row.Projeto, row.Etapa, row.Tarefa, row.Tarefa, row.UUID_Recurso,
                    row.UUID_Recurso, last_story_key, f"{row.Horas}h"
                )
                response.raise_for_status()
                task_key = response.json().get("key")
                jiralib.add_output(f"        ✅ Sucesso! Chave da Tarefa: {task_key}")
            except Exception as e:
                jiralib.add_output(f"        ❌ Falha ao criar Tarefa. Erro: {e}")
                task_key = "ERRO"
        else:
            task_key = row.TicketT
            jiralib.add_output(f"    ⏭️  Pulando Tarefa já existente: ({task_key})")
        
        # Adiciona as chaves (novas ou existentes) às listas de resultados.
        resultados["Ticket_Epico"].append(last_epic_key)
        resultados["Ticket_Estoria"].append(last_story_key)
        resultados["Ticket_Task"].append(task_key)
        
    jiralib.add_output("\n🎉 Integração com o Jira concluída!")
    return resultados

def atualizar_planilha(arquivo_excel: str, df_original: pd.DataFrame, resultados: Dict[str, List[Any]]):
    """
    Atualiza a planilha original com as chaves dos tickets criados no Jira.
    """
    jiralib.add_output("💾 Atualizando a planilha com os novos tickets...")
    try:
        wb = load_workbook(arquivo_excel)
        ws = wb[ABA_PLANILHA]
        
        # Adiciona as novas colunas ao DataFrame original
        df_original["TicketE"] = resultados["Ticket_Epico"]
        df_original["TicketS"] = resultados["Ticket_Estoria"]
        df_original["TicketT"] = resultados["Ticket_Task"]
        
        # Itera sobre as linhas do DataFrame para atualizar as células corretas.
        # Isso é mais seguro do que apagar e reescrever a planilha inteira.
        rows = dataframe_to_rows(df_original, index=False, header=True)
        
        # Obtém o cabeçalho para encontrar o índice das colunas a serem atualizadas.
        header = [cell.value for cell in ws[1]]
        col_indices = {
            "TicketE": header.index("TicketE") + 1,
            "TicketS": header.index("TicketS") + 1,
            "TicketT": header.index("TicketT") + 1
        }

        # Atualiza apenas as células necessárias, começando da segunda linha (após o cabeçalho)
        for r_idx, row in enumerate(dataframe_to_rows(df_original, index=False, header=False), 2):
            ws.cell(row=r_idx, column=col_indices["TicketE"], value=row[col_indices["TicketE"]-1])
            ws.cell(row=r_idx, column=col_indices["TicketS"], value=row[col_indices["TicketS"]-1])
            ws.cell(row=r_idx, column=col_indices["TicketT"], value=row[col_indices["TicketT"]-1])

        wb.save(arquivo_excel)
        jiralib.add_output("✔️ Planilha salva com sucesso!")

    except Exception as e:
        jiralib.add_output(f"❌ Erro Crítico ao salvar a planilha: {e}")

# --- PONTO DE ENTRADA PRINCIPAL ---
# Usar `if __name__ == "__main__":` é uma convenção padrão em Python
# que permite que o script seja importável sem executar o código principal.
if __name__ == "__main__":
    
    logging.basicConfig(filename='log.txt', level=logging.INFO)
    
    # 1. Obter o nome do arquivo da linha de comando.
    nome_arquivo = parse_argumentos()
    
    # 2. Carregar as configurações do Jira.
    # (Adicione a biblioteca jiralib ao seu ambiente)
    carregar_configuracoes_jira()

    # 3. Ler e preparar os dados da planilha.
    df_processamento = preparar_dataframe(nome_arquivo)
    
    # É uma boa ideia manter uma cópia do DataFrame original para a gravação final.
    df_original = pd.read_excel(nome_arquivo, sheet_name=ABA_PLANILHA, engine="openpyxl")

    # 4. Processar e criar as issues no Jira.
    resultados_jira = processar_issues_jira(df_processamento)

    # 5. Atualizar a planilha com os resultados.
    atualizar_planilha(nome_arquivo, df_original, resultados_jira)

    jiralib.add_output("\n🎊 Processo finalizado com sucesso! Seu trabalho foi feito! 🎊")