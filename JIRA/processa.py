import argparse
import jiralib
import json
import numpy as np
import pandas as pd

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# --- CONFIGURAÇÕES ---
# Parametros do script
parser = argparse.ArgumentParser(description="Nome da Planilha")
parser.add_argument('--nome', type=str, required=True, help='Nome da Planilha')
args = parser.parse_args()

ARQUIVO = args.nome  # Nome do arquivo Excel
ABA_PLANILHA = "Planilha"
ABA_PARAMETROS = "Parâmetros"

# --- 1. LEITURA DAS ABAS ---
df_planilha = pd.read_excel(ARQUIVO, sheet_name=ABA_PLANILHA, engine="openpyxl")
df_parametros = pd.read_excel(ARQUIVO, sheet_name=ABA_PARAMETROS, engine="openpyxl")

# --- 2. EXTRAÇÃO DA COLUNA 'API Token' ---
# Carrega setup para conexão
with open('setup.json', 'r', encoding='utf-8') as f:
    objeto = json.load(f)
    jiralib.url = objeto['url']
    jiralib.headers = objeto['headers']

# --- 3. EXTRAÇÃO E CRIAÇÃO DAS COLUNAS SOLICITADAS ---
# Seleciona as colunas necessárias
colunas_desejadas = ["Projeto",
                     "Data",
                     "Request",
                     "Task",
                     "Módulo",
                     "Tipo",
                     "Usuário Projeto",
                     "UUID PO",
                     "UUID Recurso",
                     "Épico",
                     "Estória",
                     "Tarefa",
                     "Etapa",
                     "Horas",
                     "Pontos",
                     "TicketE",
                     "TicketS",
                     "TicketT"]
df = df_planilha[colunas_desejadas].copy()

# Cria as colunas customizadas no formato solicitado
df["Request"] = df["Request"].fillna("")
df["Tipo"] = df["Tipo"].fillna("")
df["Épico"] = df["Épico"].fillna("")
df["Estória"] = df["Estória"].fillna("")
df["TicketE"] = df["TicketE"].fillna("0")
df["TicketS"] = df["TicketS"].fillna("0")
df["TicketT"] = df["TicketT"].fillna("0")
df["Nome Épico"] = "[" + df["Request"].astype(str) + "] - " + df["Épico"].astype(str)
df["Nome Estória"] = "[" + df["Tipo"].astype(str) + "] - " + df["Estória"].astype(str)

# --- 4. INTEGRAÇÃO COM JIRA ---
Epicos_criados = []
Estorias_criadas = []
Ticket_Epico = []
Ticket_Estoria = []
Ticket_Task = []
last_epic_key = None
last_story_key = None

for row in df.to_dict(orient='records'):

    EpicKey = "[" + row["Request"] + "] - " + row["Épico"]
    
    if EpicKey not in Epicos_criados:
        if row["TicketE"] == "0":
            response = jiralib.create_epic(row["Projeto"],
                                           row["Nome Épico"],
                                           row["Nome Épico"],
                                           row["Data"],
                                           row["Nome Épico"],
                                           row["Usuário Projeto"],
                                           row["Task"],
                                           row["Módulo"],
                                           row["Tipo"]
                                           )
            # Extrai a chave do épico criado
            valor_da_key = response.json().get("key")
            
            print(row["Nome Épico"])
            print(response.json())
        else:
            print(row["Nome Épico"])
            valor_da_key = row["TicketE"]
        
        last_epic_key = valor_da_key
        
        Epicos_criados.append(EpicKey)
    
    StoryKey = "[" + row["Task"] + "] - " + row["Estória"]
    
    if StoryKey not in Estorias_criadas:
        if row["TicketS"] == "0":
            response = jiralib.create_story(row["Projeto"],
                                            "História",
                                            row["Nome Estória"],
                                            row["Nome Estória"],
                                            row["UUID PO"],
                                            row["UUID Recurso"],
                                            last_epic_key,
                                            row["Pontos"]
                                            )  
            # Extrai a chave da estória criada
            valor_da_key = response.json().get("key")
            
            print("    ", row["Nome Estória"])
            print("    ", response.json())
        else:
            print("    ", row["Nome Estória"])
            valor_da_key = row["TicketS"]
        
        last_story_key = valor_da_key
        
        Estorias_criadas.append(StoryKey)
    
    if row["TicketT"] == "0":
        response = jiralib.create_task(row["Projeto"],
                                       row["Etapa"],
                                       row["Tarefa"],
                                       row["Tarefa"],
                                       row["UUID Recurso"],
                                       row["UUID Recurso"],
                                       last_story_key,
                                       str(row["Horas"]) + "h"
                                       )
        # Extrai a chave da Tarefa criada
        valor_da_key = response.json().get("key")
        
        print("        ", row["Tarefa"])
        print("        ", response.json())
    else:
        print("        ", row["Tarefa"])
        valor_da_key = row["TicketT"]
    
    Ticket_Epico.append(last_epic_key)
    Ticket_Estoria.append(last_story_key)
    Ticket_Task.append(valor_da_key)

# --- 5. SALVA AS COLUNAS NA PLANILHA ---
# Atualiza a aba "Planilha"
try:
    # Carrega o arquivo Excel com openpyxl
    wb = load_workbook(ARQUIVO)
    
    # Seleciona a aba 'Planilha'
    ws = wb[ABA_PLANILHA]
    
    # Adiciona as novas colunas ao DataFrame original, preservando as outras colunas
    df_planilha["Pontos"] = df["Pontos"]
    df_planilha["TicketE"] = Ticket_Epico
    df_planilha["TicketS"] = Ticket_Estoria
    df_planilha["TicketT"] = Ticket_Task
    
    # Limpa o conteúdo existente na aba
    ws.delete_rows(1, ws.max_row)
    
    # Reescreve o DataFrame atualizado na aba
    for r_idx, row in enumerate(dataframe_to_rows(df_planilha, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Salva o arquivo Excel
    wb.save(ARQUIVO)
    print("\nColunas salvas com sucesso na planilha.")

except Exception as e:
    print(f"\nErro ao salvar a planilha: {e}")
